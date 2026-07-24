"""AI Factory financial & engineering model generator (fully formula-driven).

Generates AI_Factory_Model.xlsx (horizon, rack architecture, location and all
other parameters are prompted at generation time) with 8 sheets:
  1. Dashboard              - headline KPIs, all formula-linked
  2. Control Panel & Inputs - every assumption in one labeled cell (blue = input)
  3. Thermal Hydraulics & MLC - ASHRAE 90.4 engine + energy/water/carbon
  4. Unit Economics & KPIs  - GPU-hour and token-level unit economics
  5. Pro Forma Financials     - leveraged model over the prompted horizon
  6. Sensitivity Engine     - 16 live scenarios (tariff x debt ratio), each with its
                              own debt schedule, GIDLR-capped tax and IRR formula
  7. Sensitivity Matrices   - IRR & DSCR grids linked to the engine (no hardcodes)

Every output cell traces back to 'Control Panel & Inputs' by formula, so changing
any input recalculates the whole workbook including both sensitivity matrices.
"""

import argparse
import sys

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FILE_NAME = "AI_Factory_Model.xlsx"

# ASHRAE climate reference library. One row per selectable location:
# (location, ASHRAE 169 climate zone, ASHRAE 90.4 max design MLC,
#  N=20yr extreme peak dry-bulb °C, coincident/peak wet-bulb °C, note).
# MLC limits and temperatures are editable reference values — verify against
# ASHRAE 90.4-2019 (Table 6.5.1.1) and ASHRAE Fundamentals climatic data
# (n=20yr return-period extremes) before relying on them for design.
LOCATION_LIBRARY = [
    ("Abu Dhabi, UAE", "1A", 0.260, 50.0, 30.0,
     "Project baseline; ASHRAE 169-2021 may reclassify as 0B"),
    ("Dubai, UAE", "0B", 0.280, 49.5, 31.0, ""),
    ("Riyadh, Saudi Arabia", "0B", 0.280, 49.0, 24.0, "Dry extreme heat"),
    ("Singapore", "0A", 0.280, 36.5, 29.5, "High wet-bulb year-round"),
    ("Mumbai, India", "0A", 0.280, 39.5, 30.0, ""),
    ("Phoenix, AZ, USA", "2B", 0.250, 48.5, 26.5, ""),
    ("Dallas, TX, USA", "3A", 0.210, 43.5, 27.5, ""),
    ("Ashburn, VA, USA", "4A", 0.190, 39.0, 27.0, "N. Virginia hub"),
    ("Chicago, IL, USA", "5A", 0.160, 38.5, 27.0, ""),
    ("Frankfurt, Germany", "5A", 0.160, 37.5, 23.0, ""),
    ("Amsterdam, Netherlands", "5A", 0.160, 35.0, 22.0, ""),
    ("Dublin, Ireland", "5A", 0.160, 28.5, 19.5, "Free-cooling friendly"),
    ("Oslo, Norway", "6A", 0.150, 32.0, 21.0, ""),
    ("Tokyo, Japan", "3A", 0.210, 37.5, 27.5, ""),
    ("Sydney, Australia", "3A", 0.210, 43.0, 24.5, ""),
]
LOCATION_INDEX = {row[0]: row for row in LOCATION_LIBRARY}
LIBRARY_SHEET = "ASHRAE Climate Library"
LIBRARY_FIRST_ROW = 5  # first data row on the library sheet

COOLING_TYPES = ("Liquid", "Air")

# Every variable parameter of the model: prompted interactively at generation
# time (Enter keeps the default). (category, name, default, unit, fmt, notes)
# Geo_Location is a menu choice from LOCATION_LIBRARY; *_Cooling_Type are
# Liquid/Air choices. Climate zone, MLC limit and N=20yr peak temperatures are
# NOT prompted — they are derived from the location via INDEX/MATCH.
PARAM_SPEC = [
    ("Geographic & Climate", "Geo_Location", "Abu Dhabi, UAE", "", None,
     "Menu choice — drives ASHRAE zone, 90.4 MLC limit and N=20yr peaks"),
    ("Geographic & Climate", "Grid_Carbon_Intensity", 0.39, "kgCO2/kWh", "0.000",
     "Assumption: Abu Dhabi grid avg (Barakah nuclear + gas + solar mix)"),
    ("Geographic & Climate", "Water_Usage_Effectiveness", 0.15, "L/kWh", "0.000",
     "Assumption: closed-loop dry coolers, minimal adiabatic assist"),
    ("Technical Specs", "VR_Rack_Count", 32, "racks", "#,##0",
     "Vera Rubin NVL72 compute racks"),
    ("Technical Specs", "VR_Rack_Power_kW", 227.0, "kW/rack", "#,##0.0",
     "Compute & scale-up busbar load per VR rack"),
    ("Technical Specs", "VR_Cooling_Type", "Liquid", "Liquid/Air", None,
     "S45 direct-to-chip liquid loop"),
    ("Technical Specs", "Network_Rack_Count", 6, "racks", "#,##0",
     "Scale-out network fabric racks"),
    ("Technical Specs", "Network_Rack_Power_kW", 100.0, "kW/rack", "#,##0.0",
     "Switching + optics load per network rack"),
    ("Technical Specs", "Network_Cooling_Type", "Air", "Liquid/Air", None,
     "Hot-aisle containment air cooling"),
    ("Technical Specs", "Storage_Rack_Count", 10, "racks", "#,##0",
     "NVMe object/block storage racks"),
    ("Technical Specs", "Storage_Rack_Power_kW", 40.0, "kW/rack", "#,##0.0",
     "Load per storage rack"),
    ("Technical Specs", "Storage_Cooling_Type", "Air", "Liquid/Air", None,
     "Hot-aisle containment air cooling"),
    ("Technical Specs", "GPUs_Per_Rack", 72, "GPUs", "#,##0",
     "NVL72 = 72 Rubin GPUs per VR compute rack"),
    ("Technical Specs", "Liquid_Cooling_Overhead", 0.05, "kW/kW", "0.0%",
     "Cooling power per kW of liquid-cooled IT (CDU pumps, dry-cooler fans)"),
    ("Technical Specs", "Air_Cooling_Overhead", 0.28, "kW/kW", "0.0%",
     "Cooling power per kW of air-cooled IT (high-ambient chillers, CRAHs)"),
    ("Technical Specs", "Liquid_Supply_Temp_FWS", 45.0, "°C", "#,##0.0",
     "S45 direct-to-chip loop"),
    ("Technical Specs", "Liquid_Return_Temp_FWR", 55.0, "°C", "#,##0.0",
     "Facility water return — sets the loop delta-T"),
    ("Technical Specs", "Coolant_Specific_Heat", 3.85, "kJ/kg·K", "0.00",
     "Assumption: PG25 water/glycol at ~50°C"),
    ("Technical Specs", "Coolant_Density", 1.03, "kg/L", "0.00",
     "Assumption: PG25 mixture density"),
    ("Technical Specs", "Target_Annualized_PUE", 1.08, "", "0.000",
     "Drives facility energy = IT energy x PUE"),
    ("Technical Specs", "Cluster_Utilization", 0.85, "%", "0.0%",
     "Assumption: avg power/capacity utilization of the fleet"),
    ("Technical Specs", "Uptime_Availability", 0.995, "%", "0.00%",
     "Assumption: contractual availability SLA"),
    ("Technical Specs", "Tokens_Per_GPU_Sec", 1800, "tok/s/GPU", "#,##0",
     "Assumption: blended inference throughput per Rubin GPU"),
    ("Technical Specs", "Token_Fleet_Share", 0.30, "%", "0.0%",
     "Share of fleet serving the API token market"),
    ("Financial & Fiscal", "Electricity_Tariff", 0.06, "$/kWh", "$#,##0.000",
     "Abu Dhabi industrial rate (Year 1)"),
    ("Financial & Fiscal", "Tariff_Escalation", 0.02, "%/yr", "0.0%",
     "Assumption: annual power price escalation"),
    ("Financial & Fiscal", "CapEx_Initial", 328000000, "USD", "$#,##0",
     "All-in benchmark midpoint: ~$91M facility (at $9.9-12.2M/MW) + 32 VR"
     " racks at ~$7.4M each (reported $6.0-8.8M). Set to ~$91M if IT is"
     " leased/vendor-financed."),
    ("Financial & Fiscal", "Residual_Value_Pct", 0.10, "%", "0.0%",
     "Assumption: pre-tax residual recovery of CapEx at end of Year 5"),
    ("Financial & Fiscal", "Debt_Ratio", 0.80, "%", "0.0%",
     "User adjustable (50% - 80%)"),
    ("Financial & Fiscal", "Interest_Rate", 0.065, "%", "0.0%",
     "Senior debt facility rate"),
    ("Financial & Fiscal", "Debt_Tenor", 5, "Years", "#,##0",
     "Annual amortization schedule"),
    ("Financial & Fiscal", "Model_Horizon_Years", 5, "Years", "#,##0",
     "Pro forma & sensitivity horizon — STRUCTURAL: regenerate the workbook"
     " to change (year columns are created at build time)"),
    ("Financial & Fiscal", "Depreciation_Life_Years", 5, "Years", "#,##0",
     "Straight-line depreciation life for the CapEx"),
    ("Financial & Fiscal", "UAE_Corporate_Tax", 0.09, "%", "0.0%",
     "Statutory corporate tax rate"),
    ("Financial & Fiscal", "GIDLR_EBITDA_Cap", 0.30, "%", "0.0%",
     "UAE interest deduction limit — applied via MIN() in the tax calc"),
    ("Financial & Fiscal", "Equity_Discount_Rate", 0.15, "%", "0.0%",
     "Assumption: hurdle rate for equity NPV"),
    ("Financial & Fiscal", "Fixed_OpEx_Annual", 10877000, "USD/yr", "$#,##0",
     "Staff, maintenance, insurance, connectivity (Year 1)"),
    ("Financial & Fiscal", "OpEx_Escalation", 0.03, "%/yr", "0.0%",
     "Assumption: annual fixed-OpEx inflation"),
    ("Revenue Strategy", "Lease_Revenue_70pct", 58800000, "USD/yr", "$#,##0",
     "70% bare-metal capacity leases (Year 1, pre-ramp)"),
    ("Revenue Strategy", "Token_Revenue_30pct", 25200000, "USD/yr", "$#,##0",
     "30% dynamic API token market (Year 1, pre-ramp)"),
    ("Revenue Strategy", "Year1_Ramp_Factor", 0.90, "%", "0.0%",
     "Assumption: Year 1 commissioning / fill-up ramp"),
    ("Revenue Strategy", "Lease_Price_Escalation", 0.00, "%/yr", "0.0%",
     "Assumption: lease repricing per year"),
    ("Revenue Strategy", "Token_Price_Decline", 0.15, "%/yr", "0.0%",
     "Core tokenomics: $/token deflation per year"),
    ("Revenue Strategy", "Token_Volume_Growth", 0.25, "%/yr", "0.0%",
     "Token demand growth (partially offsets price decline)"),
    ("Phased Expansion", "Expansion_Blocks_Per_Year", "2,2,4,4,4", "blocks/yr", None,
     "Comma-separated blocks added per year (padded/truncated to the horizon);"
     " one block = this workbook's full rack architecture. Live numeric cells"
     " are on the 'Phased Expansion' sheet."),
]

DEFAULTS = {name: default for _, name, default, *_ in PARAM_SPEC}


def _is_pct(fmt):
    return bool(fmt) and "%" in fmt


def parse_param_value(name, raw, default, fmt):
    """Parse one prompted value. Percent params accept '85%', '0.85' or '85'."""
    raw = raw.strip()
    if raw == "":
        return default
    if isinstance(default, str):
        return raw
    had_pct = raw.endswith("%")
    if had_pct:
        raw = raw[:-1].strip()
    val = float(raw.replace(",", "").replace("$", ""))
    if had_pct:
        val /= 100.0
    elif _is_pct(fmt) and val > 1.5:
        # e.g. user typed "80" for Debt_Ratio — clearly meant 80%
        print(f"    (interpreting {val:g} as {val:g}% = {val / 100:g})")
        val /= 100.0
    if isinstance(default, int) and not isinstance(default, bool):
        return int(round(val))
    return val


def parse_location(raw, default):
    """Resolve a location answer: menu number, exact or case-insensitive name."""
    raw = raw.strip()
    if raw == "":
        return default
    if raw.isdigit():
        i = int(raw)
        if 1 <= i <= len(LOCATION_LIBRARY):
            return LOCATION_LIBRARY[i - 1][0]
        print(f"    choice {i} out of range — keeping default {default!r}")
        return default
    for loc in LOCATION_INDEX:
        if loc.lower() == raw.lower() or loc.lower().startswith(raw.lower()):
            return loc
    print(f"    {raw!r} is not in the ASHRAE Climate Library — keeping"
          f" default {default!r} (add the city to LOCATION_LIBRARY or to the"
          f" library sheet to support it)")
    return default


def parse_cooling(raw, default):
    raw = raw.strip().lower()
    if raw == "":
        return default
    if raw in ("l", "liquid"):
        return "Liquid"
    if raw in ("a", "air"):
        return "Air"
    print(f"    {raw!r} is not a cooling type (Liquid/Air) — keeping"
          f" default {default!r}")
    return default


def prompt_for_params(stream=None):
    """Interactively prompt for every parameter; Enter keeps the default."""
    stream = stream or sys.stdin
    values = {}
    print("AI Factory model — parameter setup"
          " (press Enter to keep each default)\n", flush=True)
    category = None
    for cat, name, default, unit, fmt, notes in PARAM_SPEC:
        if cat != category:
            category = cat
            print(f"--- {cat} ---", flush=True)

        if name == "Geo_Location":
            print("Geo_Location — sets ASHRAE 169 zone, 90.4 MLC limit and"
                  " N=20yr peak dry/wet-bulb via the Climate Library:",
                  flush=True)
            for i, (loc, zone, mlc, db, wbt, _n) in enumerate(LOCATION_LIBRARY, 1):
                print(f"  {i:2d}) {loc}  (zone {zone}, MLC limit {mlc:.3f},"
                      f" {db:.1f}°C DB / {wbt:.1f}°C WB)", flush=True)
            print(f"  location (number or name) [{default}]: ", end="", flush=True)
            line = stream.readline()
            if line == "":
                print("(EOF — keeping defaults for remaining parameters)")
                break
            values[name] = parse_location(line, default)
            continue

        if name.endswith("_Cooling_Type"):
            print(f"{name} — {notes}", flush=True)
            print(f"  cooling type (Liquid/Air) [{default}]: ", end="", flush=True)
            line = stream.readline()
            if line == "":
                print("(EOF — keeping defaults for remaining parameters)")
                break
            values[name] = parse_cooling(line, default)
            continue

        if isinstance(default, str):
            shown = default
        elif _is_pct(fmt):
            shown = f"{default:.4g} (= {default * 100:g}%)"
        else:
            shown = f"{default:,}" if isinstance(default, int) else f"{default:g}"
        unit_s = f" [{unit}]" if unit else ""
        print(f"{name}{unit_s} — {notes}", flush=True)
        print(f"  value [{shown}]: ", end="", flush=True)
        line = stream.readline()
        if line == "":  # EOF: keep defaults for the rest
            print("(EOF — keeping defaults for remaining parameters)")
            break
        try:
            values[name] = parse_param_value(name, line, default, fmt)
        except ValueError:
            print(f"    invalid number {line.strip()!r} — keeping default {shown}")
            values[name] = default
        if values[name] < 0 if isinstance(values[name], (int, float)) else False:
            print(f"    negative value rejected — keeping default {shown}")
            values[name] = default
    out = dict(DEFAULTS)
    out.update(values)
    print("\nParameter setup complete.\n", flush=True)
    return out

FONT_FAMILY = "Segoe UI"       # labels & body
DISPLAY_FONT = "Georgia"        # banner titles (serif display)
MONO_FONT = "Consolas"          # every numeric cell (applied in polish pass)

# Fonts (financial-model color convention: blue = hardcoded input,
# black = same-sheet formula, green = link to another sheet)
TITLE_FONT = Font(name=DISPLAY_FONT, size=16, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name=DISPLAY_FONT, size=10, italic=True, color="C9D7EC")
NOTE_FONT = Font(name=FONT_FAMILY, size=9, italic=True, color="737373")
HEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
SECTION_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color="1F4E78")
BOLD_FONT = Font(name=FONT_FAMILY, size=10, bold=True)
REGULAR_FONT = Font(name=FONT_FAMILY, size=10)
INPUT_FONT = Font(name=MONO_FONT, size=10, color="0000FF")
FORMULA_FONT = Font(name=FONT_FAMILY, size=10)
LINK_FONT = Font(name=MONO_FONT, size=10, color="008000")
LINK_BOLD_FONT = Font(name=MONO_FONT, size=10, bold=True, color="008000")

NAVY_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
BAD_FILL = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")

THIN = Side(border_style="thin", color="D9D9D9")
DOUBLE = Side(border_style="double", color="1F4E78")
BORDER_DATA = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_TOTAL = Border(left=THIN, right=THIN, top=THIN, bottom=DOUBLE)

FMT_MONEY = "$#,##0"
FMT_MONEY_CT = "$#,##0.0000"
FMT_MONEY_2DP = "$#,##0.00"
FMT_PCT = "0.0%"
FMT_PCT_2DP = "0.00%"
FMT_X = '0.00"x"'
FMT_NUM = "#,##0"
FMT_NUM_1DP = "#,##0.0"
FMT_3DP = "0.000"
FMT_YEARS = '0.00" yrs"'

INPUTS_SHEET = "Control Panel & Inputs"
THERMAL_SHEET = "Thermal Hydraulics & MLC"
PROFORMA_SHEET = "Pro Forma Financials"
ENGINE_SHEET = "Sensitivity Engine"

TARIFFS = [0.04, 0.06, 0.08, 0.10]
DEBT_RATIOS = [0.50, 0.60, 0.70, 0.80]

# Row maps built as sheets are written; formulas are generated from these maps
# so cross-sheet references can never drift out of sync with the layout.
P = {}  # parameter name -> row on inputs sheet (value in col C)
T = {}  # thermal metric -> row on thermal sheet (value in col B)
F = {}  # pro forma line item -> row (Year 0..5 in cols B..G)
E = {}  # engine: scenario index -> row; also 'ebitda_grid_start'


def pref(name):
    return f"'{INPUTS_SHEET}'!$C${P[name]}"


def tref(name):
    return f"'{THERMAL_SHEET}'!$B${T[name]}"


def fref(item, col):
    return f"'{PROFORMA_SHEET}'!{col}${F[item]}"


def style_header_row(ws, row, headers):
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num, value=h)
        cell.font = HEADER_FONT
        cell.fill = NAVY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def title_block(ws, title, subtitle, width=6):
    """Navy banner across the top: Georgia title + subtitle on rows 1-2.
    `title` may be an Excel formula (leading '=') so banners stay live-linked
    to Control Panel inputs instead of baking generation-time values."""
    last = get_column_letter(width)
    for row in (1, 2):
        for col in range(1, width + 1):
            ws.cell(row=row, column=col).fill = NAVY_FILL
    ws.merge_cells(f"A1:{last}1")
    ws.merge_cells(f"A2:{last}2")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 15


# ---------------------------------------------------------------------------
# Sheet: Control Panel & Inputs
# ---------------------------------------------------------------------------
def build_inputs(wb, values):
    ws = wb.active
    ws.title = INPUTS_SHEET
    title_block(
        ws,
        '="NVIDIA VERA RUBIN NVL72 AI FACTORY — "&UPPER($C$5)',
        "Master Control Panel — edit only the yellow Value cells; every other"
        " sheet recalculates from them",
        width=5,
    )
    style_header_row(ws, 4, ["Category", "Parameter", "Value", "Unit", "Notes / Constraints"])

    n_lib = len(LOCATION_LIBRARY)
    lib_last = LIBRARY_FIRST_ROW + n_lib - 1

    def lib_lookup(col_letter):
        return (f"=INDEX('{LIBRARY_SHEET}'!${col_letter}${LIBRARY_FIRST_ROW}"
                f":${col_letter}${lib_last},"
                f"MATCH($C${P['Geo_Location']},"
                f"'{LIBRARY_SHEET}'!$A${LIBRARY_FIRST_ROW}:$A${lib_last},0))")

    # Derived climate rows inserted right after Geo_Location; values are
    # INDEX/MATCH formulas into the climate library, so retyping the location
    # cell in Excel re-derives zone, MLC limit and N=20yr peak temperatures.
    derived_after_geo = [
        ("ASHRAE_Climate_Zone", "B", "", None,
         "Derived from location — ASHRAE 169 climate zone"),
        ("ASHRAE_MLC_Limit", "C", "", "0.000",
         "Derived from location — ASHRAE 90.4 max design MLC"),
        ("Peak_Ambient_DryBulb", "D", "°C", "#,##0.0",
         "Derived from location — ASHRAE N=20yr extreme peak dry-bulb"),
        ("Coincident_WetBulb", "E", "°C", "#,##0.0",
         "Derived from location — ASHRAE N=20yr peak wet-bulb"),
    ]

    r = 5
    for cat, name, default, unit, fmt, notes in PARAM_SPEC:
        val = values.get(name, default)
        P[name] = r
        for col, v in ((1, cat), (2, name), (3, val), (4, unit), (5, notes)):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = REGULAR_FONT
            cell.border = BORDER_DATA
        vcell = ws.cell(row=r, column=3)
        vcell.font = INPUT_FONT
        vcell.fill = INPUT_FILL
        if fmt:
            vcell.number_format = fmt
        r += 1
        if name == "Geo_Location":
            for dname, lib_col, dunit, dfmt, dnotes in derived_after_geo:
                P[dname] = r
                row_vals = ((1, "Geographic & Climate"), (2, dname),
                            (3, lib_lookup(lib_col)), (4, dunit), (5, dnotes))
                for col, v in row_vals:
                    cell = ws.cell(row=r, column=col, value=v)
                    cell.font = REGULAR_FONT
                    cell.border = BORDER_DATA
                dcell = ws.cell(row=r, column=3)
                dcell.font = LINK_FONT
                if dfmt:
                    dcell.number_format = dfmt
                r += 1

    legend_row = r + 1
    ws.cell(row=legend_row, column=1,
            value="Legend: yellow cells with blue text are the editable inputs;"
                  " all other cells in this workbook are formulas.").font = NOTE_FONT
    ws.cell(row=legend_row + 1, column=1,
            value="Rows marked 'Assumption' are user-supplied estimates, not"
                  " sourced market data — revisit before investment use.").font = NOTE_FONT
    ws.cell(row=legend_row + 2, column=1,
            value="Geo_Location must exactly match a row on the 'ASHRAE Climate"
                  " Library' sheet; the four 'Derived from location' rows update"
                  " automatically. Cooling types must be 'Liquid' or 'Air'.").font = NOTE_FONT
    return ws


def build_climate_library(wb):
    ws = wb.create_sheet(title=LIBRARY_SHEET)
    title_block(
        ws,
        "ASHRAE Climate & 90.4 Compliance Library",
        "Editable reference data — verify MLC limits against ASHRAE 90.4-2019"
        " Table 6.5.1.1 and temperatures against ASHRAE Fundamentals n=20yr"
        " return-period extremes before design use",
    )
    style_header_row(ws, 4, ["Location", "ASHRAE 169 Zone",
                             "ASHRAE 90.4 Max Design MLC",
                             "N=20yr Peak Dry-Bulb (°C)",
                             "N=20yr Peak Wet-Bulb (°C)", "Notes"])
    for i, (loc, zone, mlc, db, wb_, note) in enumerate(LOCATION_LIBRARY):
        r = LIBRARY_FIRST_ROW + i
        vals = [loc, zone, mlc, db, wb_, note]
        fmts = [None, None, "0.000", "#,##0.0", "#,##0.0", None]
        for col, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = INPUT_FONT if col in (3, 4, 5) else REGULAR_FONT
            cell.border = BORDER_DATA
            if fmt:
                cell.number_format = fmt
    return ws


# ---------------------------------------------------------------------------
# Sheet: Thermal Hydraulics & MLC
# ---------------------------------------------------------------------------
def build_thermal(wb):
    ws = wb.create_sheet(title=THERMAL_SHEET)
    title_block(
        ws,
        "Thermal Hydraulics & ASHRAE 90.4 Compliance Engine",
        "S45 closed-loop fluid dynamics, 3-stage trim cooling, and annual"
        " energy / water / carbon footprint",
    )
    style_header_row(ws, 4, ["Metric / Calculation Step", "Value", "Unit", "Notes"])

    def rows():
        # (metric, formula_or_value, unit, fmt, notes)
        yield ("VR Compute Rack Load",
               f"={pref('VR_Rack_Count')}*{pref('VR_Rack_Power_kW')}",
               "kW", FMT_NUM_1DP, "VR racks x kW/rack")
        yield ("Network Rack Load",
               f"={pref('Network_Rack_Count')}*{pref('Network_Rack_Power_kW')}",
               "kW", FMT_NUM_1DP, "Network racks x kW/rack")
        yield ("Storage Rack Load",
               f"={pref('Storage_Rack_Count')}*{pref('Storage_Rack_Power_kW')}",
               "kW", FMT_NUM_1DP, "Storage racks x kW/rack")
        yield ("Total Combined IT Load", None, "kW", FMT_NUM_1DP,
               "VR compute + network + storage")
        yield ("Liquid-Cooled IT Load", None, "kW", FMT_NUM_1DP,
               "Sum of rack classes whose cooling type is 'Liquid'")
        yield ("Air-Cooled IT Load", None, "kW", FMT_NUM_1DP,
               "Total IT load minus liquid-cooled load")
        yield ("Liquid-Cooled Rack Count", None, "racks", FMT_NUM,
               "Racks on the S45 loop (cooling type = 'Liquid')")
        yield ("S45 Loop Supply Temperature",
               f"={pref('Liquid_Supply_Temp_FWS')}",
               "°C", FMT_NUM_1DP, "Facility Water Supply (FWS)")
        yield ("S45 Loop Return Temperature",
               f"={pref('Liquid_Return_Temp_FWR')}",
               "°C", FMT_NUM_1DP, "Facility Water Return (FWR)")
        yield ("S45 Loop Delta T", None, "°C", FMT_NUM_1DP,
               "Hydraulic temperature split")
        yield ("S45 Liquid Flow Rate per VR Rack", None, "LPM", FMT_NUM_1DP,
               "Physics: Q/(rho*Cp*dT)*60 — flow required to absorb one VR rack")
        yield ("Total S45 Cluster Flow Rate", None, "LPM", FMT_NUM_1DP,
               "Physics: liquid-cooled load/(rho*Cp*dT)*60")
        yield ("Peak Ambient Dry-Bulb (N=20yr)",
               f"={pref('Peak_Ambient_DryBulb')}",
               "°C", FMT_NUM_1DP, "ASHRAE n=20yr extreme — derived from location")
        yield ("Peak Wet-Bulb (N=20yr)",
               f"={pref('Coincident_WetBulb')}",
               "°C", FMT_NUM_1DP, "ASHRAE n=20yr — derived from location")
        yield ("S45 Peak Rejection Approach", None, "K", FMT_NUM_1DP,
               "FWR minus peak dry-bulb (positive = dry cooling viable)")
        yield ("Liquid Loop Cooling Power", None, "kW", FMT_NUM_1DP,
               "Liquid-cooled load x liquid cooling overhead input")
        yield ("Air Loop Cooling Power", None, "kW", FMT_NUM_1DP,
               "Air-cooled load x air cooling overhead input")
        yield ("Total Cooling Peak Power", None, "kW", FMT_NUM_1DP,
               "Combined cooling power")
        yield ("Peak Mechanical Load Component (MLC)", None, "", FMT_3DP,
               "Cooling power / total IT load")
        yield ("ASHRAE 90.4 MLC Limit (from location)", None, "", FMT_3DP,
               "Max design MLC for the selected location's climate zone")
        yield ("Safety Compliance Margin", None, "%", FMT_PCT,
               "Operational margin under the location's 90.4 limit")
        yield ("Average Utilized IT Power", None, "kW", FMT_NUM_1DP,
               "Total IT x utilization x uptime")
        yield ("Annualized PUE",
               f"={pref('Target_Annualized_PUE')}",
               "", FMT_3DP, "Facility energy multiplier on IT energy")
        yield ("Annual Facility Energy", None, "MWh/yr", FMT_NUM,
               "Avg IT power x 8760h x PUE")
        yield ("Annual Water Consumption", None, "m³/yr", FMT_NUM,
               "Facility kWh x WUE")
        yield ("Annual Carbon Emissions", None, "tCO2/yr", FMT_NUM,
               "Facility kWh x grid intensity")

    items = list(rows())
    for i, (metric, _f, unit, fmt, notes) in enumerate(items):
        T[metric] = 5 + i

    def liq(class_prefix, load_row):
        return (f'IF({pref(class_prefix + "_Cooling_Type")}="Liquid",'
                f'B{load_row},0)')

    def liq_count(class_prefix):
        return (f'IF({pref(class_prefix + "_Cooling_Type")}="Liquid",'
                f'{pref(class_prefix + "_Rack_Count")},0)')

    # Formulas that need T populated first
    formulas = {
        "Total Combined IT Load":
            f"=B{T['VR Compute Rack Load']}+B{T['Network Rack Load']}"
            f"+B{T['Storage Rack Load']}",
        "Liquid-Cooled IT Load":
            f"={liq('VR', T['VR Compute Rack Load'])}"
            f"+{liq('Network', T['Network Rack Load'])}"
            f"+{liq('Storage', T['Storage Rack Load'])}",
        "Air-Cooled IT Load":
            f"=B{T['Total Combined IT Load']}-B{T['Liquid-Cooled IT Load']}",
        "Liquid-Cooled Rack Count":
            f"={liq_count('VR')}+{liq_count('Network')}+{liq_count('Storage')}",
        "S45 Loop Delta T":
            f"=B{T['S45 Loop Return Temperature']}-B{T['S45 Loop Supply Temperature']}",
        "S45 Liquid Flow Rate per VR Rack":
            f"={pref('VR_Rack_Power_kW')}/({pref('Coolant_Density')}"
            f"*{pref('Coolant_Specific_Heat')}*B{T['S45 Loop Delta T']})*60",
        "Total S45 Cluster Flow Rate":
            f"=B{T['Liquid-Cooled IT Load']}/({pref('Coolant_Density')}"
            f"*{pref('Coolant_Specific_Heat')}*B{T['S45 Loop Delta T']})*60",
        "S45 Peak Rejection Approach":
            f"=B{T['S45 Loop Return Temperature']}"
            f"-B{T['Peak Ambient Dry-Bulb (N=20yr)']}",
        "Liquid Loop Cooling Power":
            f"=B{T['Liquid-Cooled IT Load']}*{pref('Liquid_Cooling_Overhead')}",
        "Air Loop Cooling Power":
            f"=B{T['Air-Cooled IT Load']}*{pref('Air_Cooling_Overhead')}",
        "Total Cooling Peak Power":
            f"=B{T['Liquid Loop Cooling Power']}"
            f"+B{T['Air Loop Cooling Power']}",
        "Peak Mechanical Load Component (MLC)":
            f"=B{T['Total Cooling Peak Power']}/B{T['Total Combined IT Load']}",
        "ASHRAE 90.4 MLC Limit (from location)":
            f"={pref('ASHRAE_MLC_Limit')}",
        "Safety Compliance Margin":
            f"=(B{T['ASHRAE 90.4 MLC Limit (from location)']}"
            f"-B{T['Peak Mechanical Load Component (MLC)']})"
            f"/B{T['ASHRAE 90.4 MLC Limit (from location)']}",
        "Average Utilized IT Power":
            f"=B{T['Total Combined IT Load']}*{pref('Cluster_Utilization')}"
            f"*{pref('Uptime_Availability')}",
        "Annual Facility Energy":
            f"=B{T['Average Utilized IT Power']}*8760"
            f"*B{T['Annualized PUE']}/1000",
        "Annual Water Consumption":
            f"=B{T['Annual Facility Energy']}"
            f"*{pref('Water_Usage_Effectiveness')}",
        "Annual Carbon Emissions":
            f"=B{T['Annual Facility Energy']}"
            f"*{pref('Grid_Carbon_Intensity')}",
    }

    for metric, f, unit, fmt, notes in items:
        r = T[metric]
        val = formulas.get(metric, f)
        ws.cell(row=r, column=1, value=metric).font = REGULAR_FONT
        vcell = ws.cell(row=r, column=2, value=val)
        vcell.font = BOLD_FONT if isinstance(val, str) else INPUT_FONT
        if isinstance(val, (int, float)):
            vcell.fill = INPUT_FILL
        vcell.number_format = fmt
        ws.cell(row=r, column=3, value=unit).font = REGULAR_FONT
        ws.cell(row=r, column=4, value=notes).font = REGULAR_FONT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER_DATA
    return ws


# ---------------------------------------------------------------------------
# Sheet: Pro Forma Financials (horizon = Model_Horizon_Years, set at build)
# ---------------------------------------------------------------------------
N_YEARS = 5     # overwritten by build_workbook from Model_Horizon_Years
YEAR_COLS = []  # column letters for Year 0..N, filled by build_workbook


def build_proforma(wb):
    n = N_YEARS
    last = YEAR_COLS[n]
    ws = wb.create_sheet(title=PROFORMA_SHEET)
    title_block(
        ws,
        f"{n}-Year Leveraged Financial Model & Debt Amortization",
        "UAE corporate tax (9%) with GIDLR 30%-of-EBITDA interest cap, revenue"
        " ramp & escalations, full debt schedule and equity returns",
        width=n + 2,
    )
    style_header_row(
        ws, 4, ["Line Item (USD)"] + [f"Year {t}" for t in range(n + 1)],
    )

    line_items = [
        "Revenue Ramp Factor",
        "Lease Revenue",
        "Token Revenue",
        "Gross Annual Revenue",
        "Effective Power Tariff",
        "Annual Facility Energy (kWh)",
        "Electricity Cost",
        "Fixed Non-Power OpEx",
        "Total Operational Expense (OpEx)",
        "EBITDA",
        "EBITDA Margin",
        "CapEx Initial Outlay",
        "Debt Financed Amount",
        "Equity Outlay",
        "Beginning Debt Balance",
        "Annual Debt Service (P+I)",
        "Interest Expense",
        "Debt Principal Payment",
        "Ending Debt Balance",
        "Depreciation (Straight Line)",
        "GIDLR Deductible Interest",
        "Taxable Income",
        "UAE Corporate Tax",
        "Residual Value Recovery",
        "Leveraged Free Cash Flow (FCFE)",
        "Cumulative FCFE",
        "Debt Service Coverage Ratio (DSCR)",
        "Unlevered Tax (no interest shield)",
        "Unlevered Free Cash Flow (FCFF)",
    ]
    for i, item in enumerate(line_items):
        F[item] = 5 + i

    def cells(item, y0, per_year):
        """y0: Year-0 cell; per_year(t, col): formula for Year t (1..n)."""
        r = F[item]
        vals = [y0] + [per_year(t, YEAR_COLS[t]) for t in range(1, n + 1)]
        return r, vals

    R = F  # shorthand
    life = pref("Depreciation_Life_Years")
    tenor = pref("Debt_Tenor")
    rows = []

    rows.append(cells(
        "Revenue Ramp Factor", 0,
        lambda t, c: f"={pref('Year1_Ramp_Factor')}" if t == 1 else 1,
    ))
    rows.append(cells(
        "Lease Revenue", 0,
        lambda t, c: f"={pref('Lease_Revenue_70pct')}"
                     f"*POWER(1+{pref('Lease_Price_Escalation')},{t - 1})"
                     f"*{c}{R['Revenue Ramp Factor']}",
    ))
    rows.append(cells(
        "Token Revenue", 0,
        lambda t, c: f"={pref('Token_Revenue_30pct')}"
                     f"*POWER((1+{pref('Token_Volume_Growth')})"
                     f"*(1-{pref('Token_Price_Decline')}),{t - 1})"
                     f"*{c}{R['Revenue Ramp Factor']}",
    ))
    rows.append(cells(
        "Gross Annual Revenue", 0,
        lambda t, c: f"={c}{R['Lease Revenue']}+{c}{R['Token Revenue']}",
    ))
    rows.append(cells(
        "Effective Power Tariff", 0,
        lambda t, c: f"={pref('Electricity_Tariff')}"
                     f"*POWER(1+{pref('Tariff_Escalation')},{t - 1})",
    ))
    rows.append(cells(
        "Annual Facility Energy (kWh)", 0,
        lambda t, c: f"={tref('Annual Facility Energy')}*1000",
    ))
    rows.append(cells(
        "Electricity Cost", 0,
        lambda t, c: f"={c}{R['Annual Facility Energy (kWh)']}"
                     f"*{c}{R['Effective Power Tariff']}",
    ))
    rows.append(cells(
        "Fixed Non-Power OpEx", 0,
        lambda t, c: f"={pref('Fixed_OpEx_Annual')}"
                     f"*POWER(1+{pref('OpEx_Escalation')},{t - 1})",
    ))
    rows.append(cells(
        "Total Operational Expense (OpEx)", 0,
        lambda t, c: f"={c}{R['Electricity Cost']}+{c}{R['Fixed Non-Power OpEx']}",
    ))
    rows.append(cells(
        "EBITDA", 0,
        lambda t, c: f"={c}{R['Gross Annual Revenue']}"
                     f"-{c}{R['Total Operational Expense (OpEx)']}",
    ))
    rows.append(cells(
        "EBITDA Margin", 0,
        lambda t, c: f"={c}{R['EBITDA']}/{c}{R['Gross Annual Revenue']}",
    ))
    rows.append((F["CapEx Initial Outlay"],
                 [f"={pref('CapEx_Initial')}"] + [0] * n))
    rows.append((F["Debt Financed Amount"],
                 [f"=B{R['CapEx Initial Outlay']}*{pref('Debt_Ratio')}"]
                 + [0] * n))
    rows.append((F["Equity Outlay"],
                 [f"=B{R['CapEx Initial Outlay']}-B{R['Debt Financed Amount']}"]
                 + [0] * n))
    rows.append(cells(
        "Beginning Debt Balance", 0,
        lambda t, c: (f"=B{R['Debt Financed Amount']}" if t == 1
                      else f"={YEAR_COLS[t - 1]}{R['Ending Debt Balance']}"),
    ))
    rows.append(cells(
        "Annual Debt Service (P+I)", 0,
        lambda t, c: f"=IF({t}<={tenor},"
                     f"-PMT({pref('Interest_Rate')},{tenor},"
                     f"B{R['Debt Financed Amount']}),0)",
    ))
    rows.append(cells(
        "Interest Expense", 0,
        lambda t, c: f"={c}{R['Beginning Debt Balance']}*{pref('Interest_Rate')}",
    ))
    rows.append(cells(
        "Debt Principal Payment", 0,
        lambda t, c: f"={c}{R['Annual Debt Service (P+I)']}"
                     f"-{c}{R['Interest Expense']}",
    ))
    rows.append(cells(
        "Ending Debt Balance", 0,
        lambda t, c: f"={c}{R['Beginning Debt Balance']}"
                     f"-{c}{R['Debt Principal Payment']}",
    ))
    rows.append(cells(
        "Depreciation (Straight Line)", 0,
        lambda t, c: f"=IF({t}<={life},"
                     f"B{R['CapEx Initial Outlay']}/{life},0)",
    ))
    rows.append(cells(
        "GIDLR Deductible Interest", 0,
        lambda t, c: f"=MIN({c}{R['Interest Expense']},"
                     f"{pref('GIDLR_EBITDA_Cap')}*{c}{R['EBITDA']})",
    ))
    rows.append(cells(
        "Taxable Income", 0,
        lambda t, c: f"={c}{R['EBITDA']}"
                     f"-{c}{R['Depreciation (Straight Line)']}"
                     f"-{c}{R['GIDLR Deductible Interest']}",
    ))
    rows.append(cells(
        "UAE Corporate Tax", 0,
        lambda t, c: f"=MAX(0,{c}{R['Taxable Income']})"
                     f"*{pref('UAE_Corporate_Tax')}",
    ))
    rows.append(cells(
        "Residual Value Recovery", 0,
        lambda t, c: (f"={pref('CapEx_Initial')}*{pref('Residual_Value_Pct')}"
                      if t == n else 0),
    ))
    rows.append(cells(
        "Leveraged Free Cash Flow (FCFE)",
        f"=-B{R['Equity Outlay']}",
        lambda t, c: f"={c}{R['EBITDA']}-{c}{R['Annual Debt Service (P+I)']}"
                     f"-{c}{R['UAE Corporate Tax']}"
                     f"+{c}{R['Residual Value Recovery']}",
    ))
    rows.append(cells(
        "Cumulative FCFE",
        f"=B{R['Leveraged Free Cash Flow (FCFE)']}",
        lambda t, c: f"={YEAR_COLS[t - 1]}{R['Cumulative FCFE']}"
                     f"+{c}{R['Leveraged Free Cash Flow (FCFE)']}",
    ))
    rows.append(cells(
        "Debt Service Coverage Ratio (DSCR)", 0,
        lambda t, c: f'=IF({c}{R["Annual Debt Service (P+I)"]}=0,"",'
                     f"{c}{R['EBITDA']}/{c}{R['Annual Debt Service (P+I)']})",
    ))
    rows.append(cells(
        "Unlevered Tax (no interest shield)", 0,
        lambda t, c: f"=MAX(0,{c}{R['EBITDA']}"
                     f"-{c}{R['Depreciation (Straight Line)']})"
                     f"*{pref('UAE_Corporate_Tax')}",
    ))
    rows.append(cells(
        "Unlevered Free Cash Flow (FCFF)",
        f"=-B{R['CapEx Initial Outlay']}",
        lambda t, c: f"={c}{R['EBITDA']}"
                     f"-{c}{R['Unlevered Tax (no interest shield)']}"
                     f"+{c}{R['Residual Value Recovery']}",
    ))

    pct_rows = {"Revenue Ramp Factor", "EBITDA Margin"}
    for (r, vals), item in zip(rows, line_items):
        ws.cell(row=r, column=1, value=item).font = BOLD_FONT
        for col_idx, val in enumerate(vals, start=2):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = REGULAR_FONT
            cell.border = BORDER_DATA
            if item in pct_rows:
                cell.number_format = FMT_PCT
            elif "DSCR" in item:
                cell.number_format = FMT_X
            elif item == "Effective Power Tariff":
                cell.number_format = "$#,##0.0000"
            elif item == "Annual Facility Energy (kWh)":
                cell.number_format = FMT_NUM
            else:
                cell.number_format = FMT_MONEY

    for col in range(1, n + 3):
        c = ws.cell(row=F["Leveraged Free Cash Flow (FCFE)"], column=col)
        c.fill = HIGHLIGHT_FILL
        c.font = BOLD_FONT
        c.border = BORDER_TOTAL

    # Returns summary block
    fcfe = F["Leveraged Free Cash Flow (FCFE)"]
    fcff = F["Unlevered Free Cash Flow (FCFF)"]
    cum = F["Cumulative FCFE"]
    dscr = F["Debt Service Coverage Ratio (DSCR)"]
    start = F["Unlevered Free Cash Flow (FCFF)"] + 2
    ws.cell(row=start, column=1, value="Equity Returns Summary").font = SECTION_FONT
    kpis = [
        ("Equity IRR (levered)",
         f"=IRR(B{fcfe}:{last}{fcfe})", FMT_PCT),
        ("Project IRR (unlevered)",
         f"=IRR(B{fcff}:{last}{fcff})", FMT_PCT),
        ("Equity NPV @ hurdle rate",
         f"=B{fcfe}+NPV({pref('Equity_Discount_Rate')},C{fcfe}:{last}{fcfe})",
         FMT_MONEY),
        ("MOIC (total distributions / equity)",
         f"=SUM(C{fcfe}:{last}{fcfe})/-B{fcfe}", FMT_X),
        ("Payback Period",
         f'=IFERROR(COUNTIF(B{cum}:{last}{cum},"<0")-1'
         f'+ABS(INDEX(B{cum}:{last}{cum},COUNTIF(B{cum}:{last}{cum},"<0")))'
         f'/INDEX(B{fcfe}:{last}{fcfe},COUNTIF(B{cum}:{last}{cum},"<0")+1),'
         f'">{n} yrs")',
         FMT_YEARS),
        ("Minimum DSCR", f"=MIN(C{dscr}:{last}{dscr})", FMT_X),
        ("Average DSCR", f"=AVERAGE(C{dscr}:{last}{dscr})", FMT_X),
    ]
    F["_returns_start"] = start + 1
    for i, (label, formula, fmt) in enumerate(kpis):
        r = start + 1 + i
        F[label] = r
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        cell = ws.cell(row=r, column=2, value=formula)
        cell.font = BOLD_FONT
        cell.number_format = fmt
        cell.fill = HIGHLIGHT_FILL
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = BORDER_DATA
    return ws


# ---------------------------------------------------------------------------
# Sheet: Unit Economics & KPIs
# ---------------------------------------------------------------------------
def build_unit_economics(wb):
    ws = wb.create_sheet(title="Unit Economics & KPIs")
    title_block(
        ws,
        "Unit Economics & Operational KPIs",
        "GPU-hour and token-level economics (steady-state Year 2, post-ramp)"
        " — all cells formula-driven",
    )
    style_header_row(ws, 4, ["KPI", "Value", "Unit", "Derivation"])

    U = {}

    def add(name, formula, unit, fmt, notes):
        U[name] = len(U) + 5

    kpis = [
        ("Total GPU Count",
         f"={pref('VR_Rack_Count')}*{pref('GPUs_Per_Rack')}",
         "GPUs", FMT_NUM, "Racks x GPUs per rack"),
        ("Available GPU-hours / yr", None, "GPU-hrs", FMT_NUM,
         "GPUs x 8760h x uptime"),
        ("Sold GPU-hours / yr", None, "GPU-hrs", FMT_NUM,
         "Available x utilization"),
        ("Blended Revenue per Sold GPU-hour", None, "$/GPU-hr", FMT_MONEY_2DP,
         "Year 2 gross revenue / sold GPU-hours"),
        ("Cash OpEx per Sold GPU-hour", None, "$/GPU-hr", FMT_MONEY_2DP,
         "Year 2 total OpEx / sold GPU-hours"),
        ("EBITDA per Sold GPU-hour", None, "$/GPU-hr", FMT_MONEY_2DP,
         "Revenue minus cash OpEx per GPU-hour"),
        ("CapEx per GPU", None, "$/GPU", FMT_MONEY,
         "Total CapEx / GPU count"),
        ("CapEx per MW of IT Load", None, "$/MW", FMT_MONEY,
         "Total CapEx / total IT MW"),
        ("Revenue per MW of IT Load", None, "$/MW/yr", FMT_MONEY,
         "Year 2 revenue / total IT MW"),
        ("Annual Token Capacity (token fleet)", None, "T tokens/yr", FMT_NUM_1DP,
         "Token-fleet GPUs x tok/s x seconds x utilization x uptime"),
        ("Implied Realized Price per M Tokens", None, "$/M tok", FMT_MONEY_2DP,
         "Year 2 token revenue / token capacity"),
        ("Cash Cost per M Tokens", None, "$/M tok", FMT_MONEY_2DP,
         "Token-fleet share of Year 2 OpEx / token capacity"),
        ("Energy Cost as % of Revenue", None, "%", FMT_PCT,
         "Year 2 electricity / Year 2 revenue"),
        ("Breakeven Power Tariff (EBITDA = 0)", None, "$/kWh", "$#,##0.000",
         "(Year 2 revenue - fixed OpEx) / annual kWh"),
        ("Min Revenue for 1.00x DSCR", None, "USD/yr", FMT_MONEY,
         "Debt service + Year 2 OpEx"),
        ("Revenue Headroom above 1.00x DSCR", None, "%", FMT_PCT,
         "1 - min revenue / Year 2 revenue"),
    ]
    for i, (name, *_rest) in enumerate(kpis):
        U[name] = 5 + i

    d = {
        "Available GPU-hours / yr":
            f"=B{U['Total GPU Count']}*8760*{pref('Uptime_Availability')}",
        "Sold GPU-hours / yr":
            f"=B{U['Available GPU-hours / yr']}*{pref('Cluster_Utilization')}",
        "Blended Revenue per Sold GPU-hour":
            f"={fref('Gross Annual Revenue', 'D')}/B{U['Sold GPU-hours / yr']}",
        "Cash OpEx per Sold GPU-hour":
            f"={fref('Total Operational Expense (OpEx)', 'D')}"
            f"/B{U['Sold GPU-hours / yr']}",
        "EBITDA per Sold GPU-hour":
            f"=B{U['Blended Revenue per Sold GPU-hour']}"
            f"-B{U['Cash OpEx per Sold GPU-hour']}",
        "CapEx per GPU":
            f"={pref('CapEx_Initial')}/B{U['Total GPU Count']}",
        "CapEx per MW of IT Load":
            f"={pref('CapEx_Initial')}/({tref('Total Combined IT Load')}/1000)",
        "Revenue per MW of IT Load":
            f"={fref('Gross Annual Revenue', 'D')}"
            f"/({tref('Total Combined IT Load')}/1000)",
        "Annual Token Capacity (token fleet)":
            f"=B{U['Total GPU Count']}*{pref('Token_Fleet_Share')}"
            f"*{pref('Tokens_Per_GPU_Sec')}*3600*8760"
            f"*{pref('Cluster_Utilization')}*{pref('Uptime_Availability')}"
            f"/1000000000000",
        "Implied Realized Price per M Tokens":
            f"={fref('Token Revenue', 'D')}"
            f"/(B{U['Annual Token Capacity (token fleet)']}*1000000)",
        "Cash Cost per M Tokens":
            f"={fref('Total Operational Expense (OpEx)', 'D')}"
            f"*{pref('Token_Fleet_Share')}"
            f"/(B{U['Annual Token Capacity (token fleet)']}*1000000)",
        "Energy Cost as % of Revenue":
            f"={fref('Electricity Cost', 'D')}"
            f"/{fref('Gross Annual Revenue', 'D')}",
        "Breakeven Power Tariff (EBITDA = 0)":
            f"=({fref('Gross Annual Revenue', 'D')}"
            f"-{fref('Fixed Non-Power OpEx', 'D')})"
            f"/{fref('Annual Facility Energy (kWh)', 'D')}",
        "Min Revenue for 1.00x DSCR":
            f"={fref('Annual Debt Service (P+I)', 'D')}"
            f"+{fref('Total Operational Expense (OpEx)', 'D')}",
        "Revenue Headroom above 1.00x DSCR":
            f"=1-B{U['Min Revenue for 1.00x DSCR']}"
            f"/{fref('Gross Annual Revenue', 'D')}",
    }

    for name, formula, unit, fmt, notes in kpis:
        r = U[name]
        val = d.get(name, formula)
        ws.cell(row=r, column=1, value=name).font = REGULAR_FONT
        cell = ws.cell(row=r, column=2, value=val)
        cell.font = BOLD_FONT
        cell.number_format = fmt
        ws.cell(row=r, column=3, value=unit).font = REGULAR_FONT
        ws.cell(row=r, column=4, value=notes).font = REGULAR_FONT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER_DATA

    globals()["U"] = U
    return ws


# ---------------------------------------------------------------------------
# Sheet: Sensitivity Engine (live scenarios, horizon-aware)
# ---------------------------------------------------------------------------
def build_engine(wb):
    n = N_YEARS
    nscen = len(TARIFFS) * len(DEBT_RATIOS)
    ws = wb.create_sheet(title=ENGINE_SHEET)
    title_block(
        ws,
        f"Sensitivity Engine — {nscen} Live Scenarios over {n} Years",
        "Each row re-runs the full leveraged model (debt schedule, GIDLR tax"
        " cap, residual value) for one tariff x debt-ratio combination",
        width=9 + n,
    )
    style_header_row(
        ws, 4,
        ["Scenario", "Tariff $/kWh", "Debt %", "Debt $", "Equity $",
         "Debt Service"] + [f"Y{t} FCFE" for t in range(n + 1)]
        + ["Equity IRR", "Avg DSCR"],
    )

    grid_header = 4 + nscen + 2      # EBITDA grid header row
    grid_start = grid_header + 1     # first EBITDA grid data row
    E["grid_start"] = grid_start
    irr_col = 8 + n                  # after label..DS (6 cols) + Y0..Yn
    dscr_col = 9 + n
    E["irr_col"] = irr_col
    E["dscr_col"] = dscr_col

    rate = pref("Interest_Rate")
    tenor = pref("Debt_Tenor")
    life = pref("Depreciation_Life_Years")
    capex = pref("CapEx_Initial")
    cap = pref("GIDLR_EBITDA_Cap")
    tax = pref("UAE_Corporate_Tax")
    resid = f"{pref('CapEx_Initial')}*{pref('Residual_Value_Pct')}"
    grid_cols = [get_column_letter(1 + t) for t in range(1, n + 1)]
    y_last = get_column_letter(7 + n)

    idx = 0
    for tariff in TARIFFS:
        for dr in DEBT_RATIOS:
            r = 5 + idx
            g = grid_start + idx
            E[idx] = r
            label = (f'=TEXT($B{r},"$0.00")&"/kWh @ "'
                     f'&TEXT($C{r},"0%")&" debt"')
            ws.cell(row=r, column=1, value=label).font = REGULAR_FONT
            b = ws.cell(row=r, column=2, value=tariff)
            b.font = INPUT_FONT
            b.number_format = "$#,##0.000"
            c = ws.cell(row=r, column=3, value=dr)
            c.font = INPUT_FONT
            c.number_format = "0%"
            ws.cell(row=r, column=4, value=f"={capex}*$C{r}").number_format = FMT_MONEY
            ws.cell(row=r, column=5, value=f"={capex}-$D{r}").number_format = FMT_MONEY
            ws.cell(row=r, column=6,
                    value=f"=-PMT({rate},{tenor},$D{r})").number_format = FMT_MONEY
            ws.cell(row=r, column=7, value=f"=-$E{r}").number_format = FMT_MONEY
            for t in range(1, n + 1):
                eb = f"{grid_cols[t - 1]}{g}"
                ds_t = f"IF({t}<={tenor},$F{r},0)"
                int_t = f"IF({t}<={tenor},-IPMT({rate},{t},{tenor},$D{r}),0)"
                depr_t = f"IF({t}<={life},{capex}/{life},0)"
                fcfe = (f"={eb}-{ds_t}"
                        f"-MAX(0,{eb}-{depr_t}-MIN({int_t},{cap}*{eb}))*{tax}")
                if t == n:
                    fcfe += f"+{resid}"
                cell = ws.cell(row=r, column=7 + t, value=fcfe)
                cell.number_format = FMT_MONEY
            ws.cell(row=r, column=irr_col,
                    value=f"=IRR(G{r}:{y_last}{r})").number_format = FMT_PCT
            ws.cell(row=r, column=dscr_col,
                    value=f"=AVERAGE(B{g}:{grid_cols[-1]}{g})/$F{r}"
                    ).number_format = FMT_X
            for col in range(1, dscr_col + 1):
                cell = ws.cell(row=r, column=col)
                cell.border = BORDER_DATA
                if cell.font == Font():
                    cell.font = REGULAR_FONT
            idx += 1

    # EBITDA grid: scenario-specific EBITDA per year (tariff is the only
    # scenario-dependent OpEx driver; revenue & fixed OpEx come from the pro forma)
    ws.cell(row=grid_header - 1, column=1,
            value=f"Scenario EBITDA Grid (Year 1-{n})").font = SECTION_FONT
    style_header_row(ws, grid_header,
                     ["Scenario"] + [f"Y{t} EBITDA" for t in range(1, n + 1)])
    kwh = fref("Annual Facility Energy (kWh)", "C")
    tesc = pref("Tariff_Escalation")
    for i in range(nscen):
        r = E[i]
        g = grid_start + i
        ws.cell(row=g, column=1, value=f"=A{r}").font = REGULAR_FONT
        for t in range(1, n + 1):
            col = fref("Gross Annual Revenue", YEAR_COLS[t])
            fixed = fref("Fixed Non-Power OpEx", YEAR_COLS[t])
            f = (f"={col}-{fixed}"
                 f"-{kwh}*$B{r}*POWER(1+{tesc},{t - 1})")
            cell = ws.cell(row=g, column=1 + t, value=f)
            cell.number_format = FMT_MONEY
            cell.border = BORDER_DATA
    return ws


# ---------------------------------------------------------------------------
# Sheet: Sensitivity Matrices (linked to engine)
# ---------------------------------------------------------------------------
def build_matrices(wb):
    n = N_YEARS
    irr_l = get_column_letter(E["irr_col"])
    dscr_l = get_column_letter(E["dscr_col"])
    ws = wb.create_sheet(title="Sensitivity Matrices")
    title_block(
        ws,
        f"{n}-Year Leveraged Sensitivity Matrices",
        "Power tariff ($0.04 - $0.10/kWh) vs. debt ratio (50% - 80%) — every"
        " cell is a live link to the Sensitivity Engine",
    )

    ws["A4"] = f"Matrix 1: Leveraged {n}-Year Equity IRR (%)"
    ws["A4"].font = SECTION_FONT
    headers = ["Power Tariff ($/kWh)"] + ["" for _ in DEBT_RATIOS]
    nd = len(DEBT_RATIOS)

    def axis_labels(header_row):
        # column headers live-linked to the engine's debt-ratio cells
        for di in range(nd):
            hc = ws.cell(row=header_row, column=2 + di,
                         value=f"=TEXT('{ENGINE_SHEET}'!$C${E[di]},\"0%\")"
                               f"&\" Debt\"")
            hc.font = HEADER_FONT
            hc.fill = NAVY_FILL
            hc.alignment = Alignment(horizontal="center", vertical="center")

    def matrix(header_row, col_letter, fmt):
        axis_labels(header_row)
        for ti in range(len(TARIFFS)):
            r = header_row + 1 + ti
            src_row = E[ti * nd]
            lab = ws.cell(row=r, column=1,
                          value=f"=TEXT('{ENGINE_SHEET}'!$B${src_row},"
                                f"\"$0.00\")&\" / kWh\"")
            lab.font = REGULAR_FONT
            lab.border = BORDER_DATA
            for di in range(nd):
                src = E[ti * nd + di]
                cell = ws.cell(row=r, column=2 + di,
                               value=f"='{ENGINE_SHEET}'!{col_letter}{src}")
                cell.font = LINK_FONT
                cell.border = BORDER_DATA
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")

    style_header_row(ws, 5, headers)
    matrix(5, irr_l, FMT_PCT)
    ws["A12"] = "Matrix 2: Average Debt Service Coverage Ratio (DSCR)"
    ws["A12"].font = SECTION_FONT
    style_header_row(ws, 13, headers)
    matrix(13, dscr_l, FMT_X)

    # sequential single-hue heatmaps (magnitude), plus a status rule on DSCR<1
    irr_rng = f"B6:{get_column_letter(1 + nd)}9"
    dscr_rng = f"B14:{get_column_letter(1 + nd)}17"
    scale = ColorScaleRule(start_type="min", start_color="FFFFFF",
                           end_type="max", end_color="9DC3E6")
    ws.conditional_formatting.add(irr_rng, scale)
    ws.conditional_formatting.add(
        dscr_rng, ColorScaleRule(start_type="min", start_color="FFFFFF",
                                 end_type="max", end_color="9DC3E6"))
    ws.conditional_formatting.add(
        dscr_rng,
        CellIsRule(operator="lessThan", formula=["1"], stopIfTrue=False,
                   font=Font(name=MONO_FONT, size=10, bold=True,
                             color="C00000"), fill=BAD_FILL))
    ws.conditional_formatting.add(
        irr_rng,
        CellIsRule(operator="lessThan",
                   formula=[pref("Equity_Discount_Rate")], stopIfTrue=False,
                   font=Font(name=MONO_FONT, size=10, bold=True,
                             color="C00000")))
    return ws


# ---------------------------------------------------------------------------
# Sheet: Phased Expansion (campus roll-out of identical blocks)
# ---------------------------------------------------------------------------
EXPANSION_SHEET = "Phased Expansion"
X = {}  # expansion metric -> row (years in cols B..)


def parse_expansion_schedule(raw, n):
    """'2,2,4,4,4' -> [2,2,4,4,4] padded/truncated to n years."""
    try:
        vals = [max(0, int(float(x))) for x in str(raw).split(",") if str(x).strip() != ""]
    except ValueError:
        vals = []
    if not vals:
        vals = [2, 2, 4, 4, 4]
    vals = (vals + [0] * n)[:n]
    return vals


def build_expansion(wb, values):
    n = N_YEARS
    U = globals()["U"]
    ue = "Unit Economics & KPIs"
    schedule = parse_expansion_schedule(values.get("Expansion_Blocks_Per_Year"), n)
    ws = wb.create_sheet(title=EXPANSION_SHEET)
    title_block(
        ws,
        "Phased Campus Expansion — Identical Blocks",
        "Each block replicates this workbook's rack architecture; block"
        " economics use the single-block Year-2 (steady-state) pro forma."
        " Simplification: per-vintage token-price dynamics and per-block"
        " financing are not modeled here.",
        width=n + 1,
    )
    style_header_row(ws, 4, ["Metric"] + [f"Year {t}" for t in range(1, n + 1)])

    metrics = [
        "Blocks Added (input)",
        "Cumulative Blocks",
        "Campus IT Capacity (MW)",
        "Campus GPUs",
        "Effective Revenue Blocks",
        "Campus Revenue",
        "Campus EBITDA",
        "CapEx Deployed",
        "Campus Net Cash Flow",
        "Cumulative Funding Position",
    ]
    for i, mname in enumerate(metrics):
        X[mname] = 5 + i

    block_mw = f"({tref('Total Combined IT Load')}/1000)"
    gpus = f"'{ue}'!$B${U['Total GPU Count']}"
    rev_block = fref("Gross Annual Revenue", "D")
    eb_block = fref("EBITDA", "D")
    ramp = pref("Year1_Ramp_Factor")
    capex = pref("CapEx_Initial")

    for t in range(1, n + 1):
        c = get_column_letter(1 + t)       # Year t lives in col B.. (no Year 0)
        prev = get_column_letter(t)
        cells = {
            "Blocks Added (input)": schedule[t - 1],
            "Cumulative Blocks":
                (f"={c}{X['Blocks Added (input)']}" if t == 1 else
                 f"={prev}{X['Cumulative Blocks']}+{c}{X['Blocks Added (input)']}"),
            "Campus IT Capacity (MW)":
                f"={c}{X['Cumulative Blocks']}*{block_mw}",
            "Campus GPUs":
                f"={c}{X['Cumulative Blocks']}*{gpus}",
            "Effective Revenue Blocks":
                (f"={c}{X['Blocks Added (input)']}*{ramp}" if t == 1 else
                 f"={prev}{X['Cumulative Blocks']}"
                 f"+{c}{X['Blocks Added (input)']}*{ramp}"),
            "Campus Revenue":
                f"={c}{X['Effective Revenue Blocks']}*{rev_block}",
            "Campus EBITDA":
                f"={c}{X['Effective Revenue Blocks']}*{eb_block}",
            "CapEx Deployed":
                f"={c}{X['Blocks Added (input)']}*{capex}",
            "Campus Net Cash Flow":
                f"={c}{X['Campus EBITDA']}-{c}{X['CapEx Deployed']}",
            "Cumulative Funding Position":
                (f"={c}{X['Campus Net Cash Flow']}" if t == 1 else
                 f"={prev}{X['Cumulative Funding Position']}"
                 f"+{c}{X['Campus Net Cash Flow']}"),
        }
        for mname in metrics:
            cell = ws.cell(row=X[mname], column=1 + t, value=cells[mname])
            cell.border = BORDER_DATA
            if mname == "Blocks Added (input)":
                cell.font = INPUT_FONT
                cell.fill = INPUT_FILL
                cell.number_format = FMT_NUM
            elif mname in ("Cumulative Blocks", "Campus GPUs"):
                cell.font = REGULAR_FONT
                cell.number_format = FMT_NUM
            elif mname == "Campus IT Capacity (MW)":
                cell.font = REGULAR_FONT
                cell.number_format = FMT_NUM_1DP
            elif mname == "Effective Revenue Blocks":
                cell.font = REGULAR_FONT
                cell.number_format = FMT_NUM_1DP
            else:
                cell.font = REGULAR_FONT
                cell.number_format = FMT_MONEY
    for mname in metrics:
        ws.cell(row=X[mname], column=1, value=mname).font = (
            BOLD_FONT if mname != "Blocks Added (input)" else REGULAR_FONT)
        ws.cell(row=X[mname], column=1).border = BORDER_DATA

    last = get_column_letter(1 + n)
    kpi_start = X["Cumulative Funding Position"] + 2
    ws.cell(row=kpi_start, column=1,
            value="Campus KPIs").font = SECTION_FONT
    cum = X["Cumulative Funding Position"]
    kpis = [
        ("Peak Funding Requirement",
         f"=MIN(B{cum}:{last}{cum})", FMT_MONEY),
        ("Campus Cash-Positive Year",
         f'=IF(COUNTIF(B{cum}:{last}{cum},"<0")={n},">{n}",'
         f'COUNTIF(B{cum}:{last}{cum},"<0")+1)', FMT_NUM),
        ("Final Campus IT Capacity (MW)",
         f"={last}{X['Campus IT Capacity (MW)']}", FMT_NUM_1DP),
        ("Final Campus GPUs",
         f"={last}{X['Campus GPUs']}", FMT_NUM),
        ("Run-Rate Campus EBITDA (full blocks)",
         f"={last}{X['Cumulative Blocks']}*{eb_block}", FMT_MONEY),
        ("Total Campus CapEx",
         f"=SUM(B{X['CapEx Deployed']}:{last}{X['CapEx Deployed']})", FMT_MONEY),
    ]
    for i, (label, formula, fmt) in enumerate(kpis):
        r = kpi_start + 1 + i
        X[label] = r
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        cell = ws.cell(row=r, column=2, value=formula)
        cell.font = BOLD_FONT
        cell.number_format = fmt
        cell.fill = HIGHLIGHT_FILL
        for cc in range(1, 3):
            ws.cell(row=r, column=cc).border = BORDER_DATA
    return ws


# ---------------------------------------------------------------------------
# Sheet: Dashboard
# ---------------------------------------------------------------------------
def build_dashboard(wb, values):
    """KPI stat-tile dashboard: live banner, sectioned KPI groups, status
    conditional formats, and two native charts. Row positions are exposed in
    the D map so tests never hardcode coordinates."""
    ws = wb.create_sheet(title="Dashboard", index=0)
    n = N_YEARS
    title_block(
        ws,
        f'="AI FACTORY DASHBOARD — "&TEXT({pref("VR_Rack_Count")},"0")'
        f'&"x VERA RUBIN NVL72 · "&UPPER({pref("Geo_Location")})',
        "All values live-linked; edit assumptions on 'Control Panel & Inputs'"
        " and everything on this page recalculates",
        width=11,
    )
    style_header_row(ws, 4, ["KPI", "Value", "Unit", "Source"])

    D = {}
    U = globals()["U"]
    ue = "Unit Economics & KPIs"

    sections = [
        ("RETURNS & FINANCING", [
            ("Equity IRR (levered)",
             f"='{PROFORMA_SHEET}'!B{F['Equity IRR (levered)']}", "%",
             FMT_PCT, PROFORMA_SHEET),
            ("Project IRR (unlevered)",
             f"='{PROFORMA_SHEET}'!B{F['Project IRR (unlevered)']}", "%",
             FMT_PCT, PROFORMA_SHEET),
            ("Equity NPV @ hurdle rate",
             f"='{PROFORMA_SHEET}'!B{F['Equity NPV @ hurdle rate']}", "USD",
             FMT_MONEY, PROFORMA_SHEET),
            ("MOIC",
             f"='{PROFORMA_SHEET}'!B{F['MOIC (total distributions / equity)']}",
             "x", FMT_X, PROFORMA_SHEET),
            ("Payback Period",
             f"='{PROFORMA_SHEET}'!B{F['Payback Period']}", "years",
             FMT_YEARS, PROFORMA_SHEET),
            ("Minimum DSCR",
             f"='{PROFORMA_SHEET}'!B{F['Minimum DSCR']}", "x",
             FMT_X, PROFORMA_SHEET),
        ]),
        ("OPERATIONS & COMPLIANCE", [
            ("Year 1 Gross Revenue",
             f"={fref('Gross Annual Revenue', 'C')}", "USD",
             FMT_MONEY, PROFORMA_SHEET),
            ("Year 1 EBITDA",
             f"={fref('EBITDA', 'C')}", "USD", FMT_MONEY, PROFORMA_SHEET),
            ("Year 1 EBITDA Margin",
             f"={fref('EBITDA Margin', 'C')}", "%", FMT_PCT, PROFORMA_SHEET),
            ("Peak MLC vs ASHRAE 90.4",
             f"={tref('Peak Mechanical Load Component (MLC)')}",
             "(design max below)", FMT_3DP, THERMAL_SHEET),
            ("ASHRAE 90.4 MLC Limit (location)",
             f"={pref('ASHRAE_MLC_Limit')}", "(design max)", FMT_3DP,
             INPUTS_SHEET),
            ("MLC Compliance Margin",
             f"={tref('Safety Compliance Margin')}", "%", FMT_PCT,
             THERMAL_SHEET),
            ("Location / ASHRAE 169 Zone",
             f"={pref('Geo_Location')}&\"  (zone \"&{pref('ASHRAE_Climate_Zone')}&\")\"",
             "", "@", INPUTS_SHEET),
        ]),
        ("SUSTAINABILITY", [
            ("Annual Facility Energy",
             f"={tref('Annual Facility Energy')}", "MWh/yr", FMT_NUM,
             THERMAL_SHEET),
            ("Annual Carbon Emissions",
             f"={tref('Annual Carbon Emissions')}", "tCO2/yr", FMT_NUM,
             THERMAL_SHEET),
            ("Annual Water Consumption",
             f"={tref('Annual Water Consumption')}", "m³/yr", FMT_NUM,
             THERMAL_SHEET),
        ]),
        ("UNIT ECONOMICS", [
            ("Total GPU Count", f"='{ue}'!B{U['Total GPU Count']}", "GPUs",
             FMT_NUM, ue),
            ("CapEx per GPU", f"='{ue}'!B{U['CapEx per GPU']}", "$/GPU",
             FMT_MONEY, ue),
            ("Blended Revenue per Sold GPU-hour",
             f"='{ue}'!B{U['Blended Revenue per Sold GPU-hour']}", "$/GPU-hr",
             FMT_MONEY_2DP, ue),
            ("Cash Cost per M Tokens",
             f"='{ue}'!B{U['Cash Cost per M Tokens']}", "$/M tok",
             FMT_MONEY_2DP, ue),
            ("Energy Cost as % of Revenue",
             f"='{ue}'!B{U['Energy Cost as % of Revenue']}", "%", FMT_PCT, ue),
            ("Breakeven Power Tariff",
             f"='{ue}'!B{U['Breakeven Power Tariff (EBITDA = 0)']}", "$/kWh",
             "$#,##0.000", ue),
        ]),
        ("CAMPUS EXPANSION", [
            ("Final Campus IT Capacity",
             f"='{EXPANSION_SHEET}'!B{X['Final Campus IT Capacity (MW)']}",
             "MW", FMT_NUM_1DP, EXPANSION_SHEET),
            ("Final Campus GPUs",
             f"='{EXPANSION_SHEET}'!B{X['Final Campus GPUs']}", "GPUs",
             FMT_NUM, EXPANSION_SHEET),
            ("Total Campus CapEx",
             f"='{EXPANSION_SHEET}'!B{X['Total Campus CapEx']}", "USD",
             FMT_MONEY, EXPANSION_SHEET),
            ("Peak Funding Requirement",
             f"='{EXPANSION_SHEET}'!B{X['Peak Funding Requirement']}", "USD",
             FMT_MONEY, EXPANSION_SHEET),
            ("Run-Rate Campus EBITDA",
             f"='{EXPANSION_SHEET}'!B{X['Run-Rate Campus EBITDA (full blocks)']}",
             "USD/yr", FMT_MONEY, EXPANSION_SHEET),
        ]),
    ]

    r = 5
    for section, kpis in sections:
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.fill = SECTION_FILL
            cell.border = BORDER_DATA
        ws.cell(row=r, column=1, value=section).font = SECTION_FONT
        ws.row_dimensions[r].height = 16
        r += 1
        for name, formula, unit, fmt, source in kpis:
            D[name] = r
            ws.cell(row=r, column=1, value=name).font = REGULAR_FONT
            cell = ws.cell(row=r, column=2, value=formula)
            cell.font = LINK_BOLD_FONT
            cell.number_format = fmt
            ws.cell(row=r, column=3, value=unit).font = NOTE_FONT
            ws.cell(row=r, column=4, value=source).font = NOTE_FONT
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = BORDER_DATA
            r += 1

    # Status conditional formats (red = breach, needs attention)
    bad_font = Font(name=MONO_FONT, size=10, bold=True, color="C00000")
    ws.conditional_formatting.add(
        f"B{D['Minimum DSCR']}",
        CellIsRule(operator="lessThan", formula=["1"], font=bad_font,
                   fill=BAD_FILL))
    ws.conditional_formatting.add(
        f"B{D['Equity IRR (levered)']}",
        CellIsRule(operator="lessThan",
                   formula=[pref("Equity_Discount_Rate")], font=bad_font))
    ws.conditional_formatting.add(
        f"B{D['Equity NPV @ hurdle rate']}",
        CellIsRule(operator="lessThan", formula=["0"], font=bad_font))
    ws.conditional_formatting.add(
        f"B{D['Peak MLC vs ASHRAE 90.4']}",
        CellIsRule(operator="greaterThan",
                   formula=[pref("ASHRAE_MLC_Limit")], font=bad_font,
                   fill=BAD_FILL))

    # Charts (single series each — the title names the series, no legend)
    pro = wb[PROFORMA_SHEET]
    fcfe_row = F["Leveraged Free Cash Flow (FCFE)"]
    bar = BarChart()
    bar.type = "col"
    bar.title = "Leveraged FCFE by Year (USD)"
    bar.legend = None
    bar.height, bar.width = 6.6, 10.6
    data = Reference(pro, min_col=2, max_col=2 + n,
                     min_row=fcfe_row, max_row=fcfe_row)
    cats = Reference(pro, min_col=2, max_col=2 + n, min_row=4, max_row=4)
    bar.add_data(data, titles_from_data=False, from_rows=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties = GraphicalProperties(solidFill="1F4E78")
    bar.gapWidth = 60
    ws.add_chart(bar, "F5")

    xs = wb[EXPANSION_SHEET]
    line = LineChart()
    line.title = "Campus Expansion — Cumulative Funding Position (USD)"
    line.legend = None
    line.height, line.width = 6.6, 10.6
    ldata = Reference(xs, min_col=2, max_col=1 + n,
                      min_row=X["Cumulative Funding Position"],
                      max_row=X["Cumulative Funding Position"])
    lcats = Reference(xs, min_col=2, max_col=1 + n, min_row=4, max_row=4)
    line.add_data(ldata, titles_from_data=False, from_rows=True)
    line.set_categories(lcats)
    line.series[0].graphicalProperties = GraphicalProperties(
        ln=LineProperties(solidFill="C55A11", w=28000))
    ws.add_chart(line, "F19")

    globals()["D"] = D
    return ws


TAB_COLORS = {
    "Dashboard": "1F4E78",
    INPUTS_SHEET: "FFC000",
    LIBRARY_SHEET: "A6A6A6",
    THERMAL_SHEET: "5B9BD5",
    PROFORMA_SHEET: "70AD47",
    "Unit Economics & KPIs": "ED7D31",
    ENGINE_SHEET: "7030A0",
    "Sensitivity Matrices": "C00000",
    EXPANSION_SHEET: "00B0A0",
}

ZEBRA_SHEETS = {INPUTS_SHEET, LIBRARY_SHEET, THERMAL_SHEET,
                "Unit Economics & KPIs"}


def polish(wb):
    """Final visual pass: hide gridlines, freeze headers, color tabs, zebra
    banding on reference tables, and Consolas for every numeric cell."""
    numeric_chars = set("0#%$")
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A5"
        if ws.title in TAB_COLORS:
            ws.sheet_properties.tabColor = TAB_COLORS[ws.title]
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                f = cell.font
                if (cell.value is not None
                        and any(ch in numeric_chars
                                for ch in (cell.number_format or ""))
                        and f.name not in (MONO_FONT, DISPLAY_FONT)):
                    cell.font = Font(name=MONO_FONT, size=f.size or 10,
                                     bold=f.bold, italic=f.italic,
                                     color=f.color)
                if (ws.title in ZEBRA_SHEETS and cell.row >= 5
                        and cell.row % 2 == 1
                        and cell.border.left.style is not None
                        and (cell.fill is None
                             or cell.fill.fill_type is None)):
                    cell.fill = ZEBRA_FILL


def autosize(wb):
    from openpyxl.utils import get_column_letter
    for ws in wb.worksheets:
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                if cell.value is None or cell.row in (1, 2):
                    continue
                v = str(cell.value)
                if v.startswith("="):  # formula text length is meaningless
                    v = "0" * 14
                max_len = max(max_len, len(v))
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 44)


def build_workbook(values=None):
    global N_YEARS
    values = dict(DEFAULTS, **(values or {}))
    n = max(2, min(int(values.get("Model_Horizon_Years", 5)), 15))
    values["Model_Horizon_Years"] = n
    N_YEARS = n
    YEAR_COLS[:] = [get_column_letter(2 + t) for t in range(n + 1)]
    P.clear(); T.clear(); F.clear(); E.clear(); X.clear()
    wb = openpyxl.Workbook()
    build_inputs(wb, values)
    build_climate_library(wb)
    build_thermal(wb)
    build_proforma(wb)
    build_unit_economics(wb)
    build_engine(wb)
    build_matrices(wb)
    build_expansion(wb, values)
    build_dashboard(wb, values)
    autosize(wb)
    polish(wb)
    return wb, {"P": dict(P), "T": dict(T), "F": dict(F), "E": dict(E),
                "U": dict(globals()["U"]), "X": dict(X),
                "D": dict(globals()["D"])}


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Generate the AI Factory Excel model. By default every"
                    " variable parameter is prompted interactively"
                    " (Enter keeps the default).")
    parser.add_argument("--defaults", action="store_true",
                        help="skip all prompts and use the built-in defaults")
    parser.add_argument("--interactive", action="store_true",
                        help="force prompting even when stdin is not a TTY"
                             " (answers can be piped, one per line)")
    parser.add_argument("--output", default=FILE_NAME,
                        help=f"output .xlsx path (default: {FILE_NAME})")
    args = parser.parse_args(argv)

    interactive = args.interactive or (not args.defaults and sys.stdin.isatty())
    if interactive:
        values = prompt_for_params()
    else:
        if not args.defaults:
            print("stdin is not a TTY — using built-in defaults"
                  " (pass --interactive to pipe answers)")
        values = dict(DEFAULTS)

    wb, _maps = build_workbook(values)
    wb.save(args.output)
    print(f"Successfully generated financial & engineering model: {args.output}")


if __name__ == "__main__":
    main()
