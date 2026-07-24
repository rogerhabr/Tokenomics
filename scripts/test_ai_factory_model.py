"""Verification harness for AI_Factory_Model.xlsx.

Re-implements the entire model independently in pure Python (no openpyxl
formula evaluation involved) and compares key cells of the recalculated
workbook against the expected values, including all 16 sensitivity-engine
IRRs. Works for any model horizon (Model_Horizon_Years); pass a workbook
path and horizon on the command line to verify a non-default build:

    python3 test_ai_factory_model.py [workbook.xlsx] [horizon]

Run scripts/create_ai_factory_model.py and the xlsx recalc step first.
Exit code 0 = all checks pass.
"""

import sys

import openpyxl
from openpyxl.utils import get_column_letter

from create_ai_factory_model import (
    DEBT_RATIOS,
    FILE_NAME,
    LOCATION_INDEX,
    PROFORMA_SHEET,
    TARIFFS,
    YEAR_COLS,
    build_workbook,
)

# ---------------------------------------------------------------------------
# Independent model implementation
# ---------------------------------------------------------------------------
INP = dict(
    location="Abu Dhabi, UAE",
    vr_count=32, vr_kw=227.0, vr_cool="Liquid",
    net_count=6, net_kw=100.0, net_cool="Air",
    sto_count=10, sto_kw=40.0, sto_cool="Air",
    liq_oh=0.05, air_oh=0.28,
    fws=45.0, fwr=55.0, cp=3.85, rho=1.03,
    gpus_per_rack=72,
    pue=1.08, util=0.85, uptime=0.995, tokens_gpu_s=1800, token_share=0.30,
    tariff=0.06, tariff_esc=0.02, capex=328_000_000, resid_pct=0.10,
    blocks_per_year=(2, 2, 4, 4, 4),
    debt_ratio=0.80, rate=0.065, tenor=5, tax=0.09, gidlr=0.30,
    horizon=5, dep_life=5,
    hurdle=0.15, fixed_opex=10_877_000, opex_esc=0.03,
    lease_rev=58_800_000, token_rev=25_200_000, ramp=0.90,
    lease_esc=0.00, tok_decline=0.15, tok_growth=0.25,
    wue=0.15, carbon=0.39,
)


def pmt(rate, n, pv):
    return pv * rate / (1 - (1 + rate) ** -n)


def irr(cashflows, lo=-0.99, hi=10.0, tol=1e-10):
    def npv(r):
        return sum(cf / (1 + r) ** t for t, cf in enumerate(cashflows))
    f_lo, f_hi = npv(lo), npv(hi)
    if f_lo * f_hi > 0:
        raise ValueError("IRR not bracketed")
    for _ in range(200):
        mid = (lo + hi) / 2
        f_mid = npv(mid)
        if abs(f_mid) < tol:
            return mid
        if f_lo * f_mid < 0:
            hi = mid
        else:
            lo, f_lo = mid, f_mid
    return (lo + hi) / 2


def debt_schedule(debt, rate, tenor, n):
    """Per-year (debt service, interest, ending balance) over n years."""
    ds_full = pmt(rate, tenor, debt)
    bal, out = debt, []
    for y in range(1, n + 1):
        ds = ds_full if y <= tenor else 0.0
        i = bal * rate
        bal -= ds - i
        out.append((ds, i, bal))
    return ds_full, out


def compute_model(p):
    n = p["horizon"]
    life = p["dep_life"]
    m = {"n": n}
    loads = {
        "vr": p["vr_count"] * p["vr_kw"],
        "net": p["net_count"] * p["net_kw"],
        "sto": p["sto_count"] * p["sto_kw"],
    }
    it_kw = sum(loads.values())
    m["it_kw"] = it_kw
    liquid_kw = sum(loads[k] for k in loads if p[f"{k}_cool"] == "Liquid")
    air_kw = it_kw - liquid_kw
    m["liquid_kw"] = liquid_kw
    m["liquid_racks"] = sum(
        p[f"{k}_count"] for k in ("vr", "net", "sto")
        if p[f"{k}_cool"] == "Liquid")
    dt = p["fwr"] - p["fws"]
    m["flow_per_rack"] = p["vr_kw"] / (p["rho"] * p["cp"] * dt) * 60
    m["flow_lpm"] = liquid_kw / (p["rho"] * p["cp"] * dt) * 60
    m["mlc"] = (liquid_kw * p["liq_oh"] + air_kw * p["air_oh"]) / it_kw
    loc = LOCATION_INDEX[p["location"]]
    m["zone"], m["mlc_limit"], m["db20"], m["wb20"] = loc[1], loc[2], loc[3], loc[4]
    m["mlc_margin"] = (m["mlc_limit"] - m["mlc"]) / m["mlc_limit"]
    avg_it_kw = it_kw * p["util"] * p["uptime"]
    m["energy_mwh"] = avg_it_kw * 8760 * p["pue"] / 1000
    kwh = m["energy_mwh"] * 1000
    m["kwh"] = kwh
    m["water_m3"] = m["energy_mwh"] * p["wue"]
    m["co2_t"] = m["energy_mwh"] * p["carbon"]

    # Revenue / OpEx / EBITDA per year (index 0 = Year 1)
    ramp = [p["ramp"]] + [1] * (n - 1)
    lease = [p["lease_rev"] * (1 + p["lease_esc"]) ** t * ramp[t]
             for t in range(n)]
    tok_f = (1 + p["tok_growth"]) * (1 - p["tok_decline"])
    token = [p["token_rev"] * tok_f ** t * ramp[t] for t in range(n)]
    rev = [a + b for a, b in zip(lease, token)]
    elec = [kwh * p["tariff"] * (1 + p["tariff_esc"]) ** t for t in range(n)]
    fixed = [p["fixed_opex"] * (1 + p["opex_esc"]) ** t for t in range(n)]
    opex = [a + b for a, b in zip(elec, fixed)]
    ebitda = [r - o for r, o in zip(rev, opex)]
    m.update(rev=rev, token=token, elec=elec, opex=opex, ebitda=ebitda)

    debt = p["capex"] * p["debt_ratio"]
    equity = p["capex"] - debt
    ds_full, sched = debt_schedule(debt, p["rate"], p["tenor"], n)
    ds = [s[0] for s in sched]
    interest = [s[1] for s in sched]
    ending = [s[2] for s in sched]
    m.update(debt=debt, equity=equity, ds=ds_full, ending=ending)

    depr = [p["capex"] / life if y <= life else 0.0 for y in range(1, n + 1)]
    resid = p["capex"] * p["resid_pct"]
    ded_int = [min(i, p["gidlr"] * e) for i, e in zip(interest, ebitda)]
    taxable = [e - d - di for e, d, di in zip(ebitda, depr, ded_int)]
    tax = [max(0, ti) * p["tax"] for ti in taxable]
    fcfe = [e - s - tx for e, s, tx in zip(ebitda, ds, tax)]
    fcfe[n - 1] += resid
    m.update(tax=tax, fcfe=[-equity] + fcfe)

    cum, running = [], -equity
    cum.append(running)
    for cf in fcfe:
        running += cf
        cum.append(running)
    m["cum"] = cum
    m["equity_irr"] = irr(m["fcfe"])
    m["npv"] = sum(cf / (1 + p["hurdle"]) ** t for t, cf in enumerate(m["fcfe"]))
    m["moic"] = sum(m["fcfe"][1:]) / equity
    n_neg = sum(1 for c in cum if c < 0)
    if n_neg >= len(cum):
        m["payback"] = None  # never pays back within the horizon
    else:
        m["payback"] = (n_neg - 1) + abs(cum[n_neg - 1]) / m["fcfe"][n_neg]
    dscr = [e / s for e, s in zip(ebitda, ds) if s > 0]
    m["dscr"] = dscr
    m["min_dscr"] = min(dscr)
    m["avg_dscr"] = sum(dscr) / len(dscr)

    utax = [max(0, e - d) * p["tax"] for e, d in zip(ebitda, depr)]
    fcff = [e - t for e, t in zip(ebitda, utax)]
    fcff[n - 1] += resid
    m["fcff"] = [-p["capex"]] + fcff
    m["project_irr"] = irr(m["fcff"])

    # Unit economics (Year 2, post-ramp)
    gpus = p["vr_count"] * p["gpus_per_rack"]
    avail = gpus * 8760 * p["uptime"]
    sold = avail * p["util"]
    m["gpus"] = gpus
    m["rev_gpu_hr"] = rev[1] / sold
    m["opex_gpu_hr"] = opex[1] / sold
    m["capex_gpu"] = p["capex"] / gpus
    tok_cap_t = (gpus * p["token_share"] * p["tokens_gpu_s"] * 3600 * 8760
                 * p["util"] * p["uptime"] / 1e12)
    m["tok_cap_t"] = tok_cap_t
    m["price_mtok"] = token[1] / (tok_cap_t * 1e6)
    m["cost_mtok"] = opex[1] * p["token_share"] / (tok_cap_t * 1e6)
    m["energy_pct"] = elec[1] / rev[1]
    m["breakeven_tariff"] = (rev[1] - fixed[1]) / kwh
    m["min_rev_dscr1"] = ds_full + opex[1]

    # Sensitivity engine: same revenue/fixed-opex, scenario tariff & leverage
    m["engine"] = {}
    for tariff in TARIFFS:
        for dr in DEBT_RATIOS:
            sd = p["capex"] * dr
            se = p["capex"] - sd
            s_full, s_sched = debt_schedule(sd, p["rate"], p["tenor"], n)
            s_ds = [x[0] for x in s_sched]
            s_int = [x[1] for x in s_sched]
            s_elec = [kwh * tariff * (1 + p["tariff_esc"]) ** t
                      for t in range(n)]
            s_eb = [r - f - el for r, f, el in zip(rev, fixed, s_elec)]
            s_ded = [min(i, p["gidlr"] * e) for i, e in zip(s_int, s_eb)]
            s_tax = [max(0, e - d - di) * p["tax"]
                     for e, d, di in zip(s_eb, depr, s_ded)]
            s_fcfe = [e - s - tx for e, s, tx in zip(s_eb, s_ds, s_tax)]
            s_fcfe[n - 1] += resid
            m["engine"][(tariff, dr)] = dict(
                irr=irr([-se] + s_fcfe),
                avg_dscr=sum(s_eb) / n / s_full,
            )

    # Phased expansion (identical blocks; block economics = Year-2 steady state)
    sched = (list(p.get("blocks_per_year", ())) + [0] * n)[:n]
    cum_b, eff, cume, cash = 0, [], [], 0.0
    exp = {"added": sched, "cum_blocks": [], "eb": [], "cum_fund": []}
    for t in range(n):
        eff_t = cum_b + sched[t] * p["ramp"]
        cum_b += sched[t]
        eb_t = eff_t * ebitda[1]
        cash += eb_t - sched[t] * p["capex"]
        exp["cum_blocks"].append(cum_b)
        exp["eb"].append(eb_t)
        exp["cum_fund"].append(cash)
    exp["final_mw"] = cum_b * it_kw / 1000
    exp["final_gpus"] = cum_b * gpus
    exp["peak_funding"] = min(exp["cum_fund"])
    exp["runrate_ebitda"] = cum_b * ebitda[1]
    exp["capex_total"] = sum(sched) * p["capex"]
    n_neg_x = sum(1 for c in exp["cum_fund"] if c < 0)
    exp["cash_positive_year"] = f">{n}" if n_neg_x == n else n_neg_x + 1
    m["expansion"] = exp
    return m


# ---------------------------------------------------------------------------
# Comparison
# ---------------------------------------------------------------------------
def main():
    file_name = sys.argv[1] if len(sys.argv) > 1 else FILE_NAME
    horizon = int(sys.argv[2]) if len(sys.argv) > 2 else 5

    _, maps = build_workbook({"Model_Horizon_Years": horizon})  # coords only
    P, T, F, U, E, D = (maps[k] for k in ("P", "T", "F", "U", "E", "D"))
    m = compute_model(dict(INP, horizon=horizon))
    n = m["n"]

    wb = openpyxl.load_workbook(file_name, data_only=True)
    thermal = wb["Thermal Hydraulics & MLC"]
    pro = wb[PROFORMA_SHEET]
    ue = wb["Unit Economics & KPIs"]
    eng = wb["Sensitivity Engine"]
    dash = wb["Dashboard"]
    mat = wb["Sensitivity Matrices"]

    checks = []

    def chk(label, actual, expected, tol=1e-4):
        if isinstance(expected, str):
            checks.append((label, actual, expected, actual == expected))
            return
        if actual is None:
            checks.append((label, "MISSING (None)", expected, False))
            return
        rel = abs(actual - expected) / max(abs(expected), 1e-9)
        checks.append((label, actual, expected, rel <= tol))

    chk("Thermal: total IT kW", thermal[f"B{T['Total Combined IT Load']}"].value,
        m["it_kw"])
    chk("Thermal: liquid-cooled kW",
        thermal[f"B{T['Liquid-Cooled IT Load']}"].value, m["liquid_kw"])
    chk("Thermal: S45 flow LPM",
        thermal[f"B{T['Total S45 Cluster Flow Rate']}"].value, m["flow_lpm"])
    chk("Thermal: S45 flow per VR rack",
        thermal[f"B{T['S45 Liquid Flow Rate per VR Rack']}"].value,
        m["flow_per_rack"])
    chk("Thermal: peak MLC",
        thermal[f"B{T['Peak Mechanical Load Component (MLC)']}"].value, m["mlc"])
    chk("Thermal: MLC limit (location)",
        thermal[f"B{T['ASHRAE 90.4 MLC Limit (from location)']}"].value,
        m["mlc_limit"])
    chk("Thermal: N=20yr peak DB (location)",
        thermal[f"B{T['Peak Ambient Dry-Bulb (N=20yr)']}"].value, m["db20"])
    chk("Thermal: MLC margin",
        thermal[f"B{T['Safety Compliance Margin']}"].value, m["mlc_margin"])
    chk("Thermal: energy MWh",
        thermal[f"B{T['Annual Facility Energy']}"].value, m["energy_mwh"])
    chk("Thermal: water m3",
        thermal[f"B{T['Annual Water Consumption']}"].value, m["water_m3"])
    chk("Thermal: CO2 t",
        thermal[f"B{T['Annual Carbon Emissions']}"].value, m["co2_t"])

    for t in range(1, n + 1):
        col = YEAR_COLS[t]
        chk(f"ProForma: Y{t} revenue",
            pro[f"{col}{F['Gross Annual Revenue']}"].value, m["rev"][t - 1])
        chk(f"ProForma: Y{t} EBITDA",
            pro[f"{col}{F['EBITDA']}"].value, m["ebitda"][t - 1])
        chk(f"ProForma: Y{t} FCFE",
            pro[f"{col}{F['Leveraged Free Cash Flow (FCFE)']}"].value,
            m["fcfe"][t])

    chk("ProForma: Y0 FCFE (=-equity)",
        pro[f"B{F['Leveraged Free Cash Flow (FCFE)']}"].value, m["fcfe"][0])
    chk("ProForma: debt service Y1",
        pro[f"C{F['Annual Debt Service (P+I)']}"].value, m["ds"])
    tenor_end = min(INP["tenor"], n)
    chk(f"ProForma: Y{tenor_end} ending debt ~ 0",
        pro[f"{YEAR_COLS[tenor_end]}{F['Ending Debt Balance']}"].value + 1,
        m["ending"][tenor_end - 1] + 1, tol=1e-6)
    chk("ProForma: equity IRR",
        pro[f"B{F['Equity IRR (levered)']}"].value, m["equity_irr"], tol=1e-5)
    chk("ProForma: project IRR",
        pro[f"B{F['Project IRR (unlevered)']}"].value, m["project_irr"],
        tol=1e-5)
    chk("ProForma: NPV", pro[f"B{F['Equity NPV @ hurdle rate']}"].value,
        m["npv"])
    chk("ProForma: MOIC",
        pro[f"B{F['MOIC (total distributions / equity)']}"].value, m["moic"])
    if m["payback"] is None:
        pb = pro[f"B{F['Payback Period']}"].value
        chk("ProForma: payback (text)", 1 if str(pb).startswith(">") else 0, 1,
            tol=0)
    else:
        chk("ProForma: payback", pro[f"B{F['Payback Period']}"].value,
            m["payback"])
    chk("ProForma: min DSCR", pro[f"B{F['Minimum DSCR']}"].value,
        m["min_dscr"])
    chk("ProForma: avg DSCR", pro[f"B{F['Average DSCR']}"].value,
        m["avg_dscr"])

    chk("UnitEcon: GPUs", ue[f"B{U['Total GPU Count']}"].value, m["gpus"])
    chk("UnitEcon: rev/GPU-hr",
        ue[f"B{U['Blended Revenue per Sold GPU-hour']}"].value,
        m["rev_gpu_hr"])
    chk("UnitEcon: capex/GPU", ue[f"B{U['CapEx per GPU']}"].value,
        m["capex_gpu"])
    chk("UnitEcon: token capacity (T)",
        ue[f"B{U['Annual Token Capacity (token fleet)']}"].value,
        m["tok_cap_t"])
    chk("UnitEcon: $/M tok realized",
        ue[f"B{U['Implied Realized Price per M Tokens']}"].value,
        m["price_mtok"])
    chk("UnitEcon: cost/M tok", ue[f"B{U['Cash Cost per M Tokens']}"].value,
        m["cost_mtok"])
    chk("UnitEcon: energy % rev",
        ue[f"B{U['Energy Cost as % of Revenue']}"].value, m["energy_pct"])
    chk("UnitEcon: breakeven tariff",
        ue[f"B{U['Breakeven Power Tariff (EBITDA = 0)']}"].value,
        m["breakeven_tariff"])

    irr_l = get_column_letter(E["irr_col"])
    dscr_l = get_column_letter(E["dscr_col"])
    for ti, tariff in enumerate(TARIFFS):
        for di, dr in enumerate(DEBT_RATIOS):
            row = E[ti * len(DEBT_RATIOS) + di]
            exp = m["engine"][(tariff, dr)]
            chk(f"Engine IRR: ${tariff:.2f}/{int(dr*100)}%",
                eng[f"{irr_l}{row}"].value, exp["irr"], tol=1e-5)
            chk(f"Engine DSCR: ${tariff:.2f}/{int(dr*100)}%",
                eng[f"{dscr_l}{row}"].value, exp["avg_dscr"])
            chk(f"Matrix IRR: ${tariff:.2f}/{int(dr*100)}%",
                mat.cell(row=6 + ti, column=2 + di).value, exp["irr"],
                tol=1e-5)

    chk("Dashboard: equity IRR",
        dash[f"B{D['Equity IRR (levered)']}"].value, m["equity_irr"],
        tol=1e-5)
    chk("Dashboard: CO2",
        dash[f"B{D['Annual Carbon Emissions']}"].value, m["co2_t"])
    chk("Dashboard: location text",
        dash[f"B{D['Location / ASHRAE 169 Zone']}"].value,
        f"{INP['location']}  (zone {m['zone']})")
    chk("Dashboard: peak funding link",
        dash[f"B{D['Peak Funding Requirement']}"].value,
        m["expansion"]["peak_funding"])

    Xm = maps["X"]
    exp = m["expansion"]
    xs = wb["Phased Expansion"]
    for t in range(1, n + 1):
        xc = get_column_letter(1 + t)
        chk(f"Expansion: Y{t} cum funding",
            xs[f"{xc}{Xm['Cumulative Funding Position']}"].value,
            exp["cum_fund"][t - 1])
    chk("Expansion: final MW",
        xs[f"B{Xm['Final Campus IT Capacity (MW)']}"].value, exp["final_mw"])
    chk("Expansion: final GPUs",
        xs[f"B{Xm['Final Campus GPUs']}"].value, exp["final_gpus"])
    chk("Expansion: peak funding",
        xs[f"B{Xm['Peak Funding Requirement']}"].value, exp["peak_funding"])
    chk("Expansion: run-rate EBITDA",
        xs[f"B{Xm['Run-Rate Campus EBITDA (full blocks)']}"].value,
        exp["runrate_ebitda"])
    chk("Expansion: total campus CapEx",
        xs[f"B{Xm['Total Campus CapEx']}"].value, exp["capex_total"])
    cp = xs[f"B{Xm['Campus Cash-Positive Year']}"].value
    chk("Expansion: cash-positive year",
        1 if str(cp) == str(exp["cash_positive_year"]) else 0, 1, tol=0)

    failures = [c for c in checks if not c[3]]
    for label, actual, expected, ok in checks:
        if not ok:
            print(f"[FAIL] {label}: got {actual!r}, expected {expected!r}")
    print(f"\n{len(checks) - len(failures)}/{len(checks)} checks passed"
          f" (horizon = {n} years, file = {file_name}).")
    if failures:
        sys.exit(1)
    print("All checks passed — workbook is fully verified.")


if __name__ == "__main__":
    main()
